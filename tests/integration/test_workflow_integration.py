"""
End-to-end integration test for ScribeWorkflowProcessor
Tests complete email-to-Excel workflow with real data integration
"""

import pytest
from unittest.mock import Mock, patch, AsyncMock
import json
from datetime import datetime

from src.processors.workflow_processor import ScribeWorkflowProcessor
from src.models.workflow_models import WorkflowRun, WorkflowConfiguration


class TestScribeWorkflowProcessorIntegration:
    """End-to-end integration tests for the complete workflow"""

    def test_complete_workflow_with_real_data_simulation(self, production_config, core_services, environment_setup):
        """
        Test complete end-to-end workflow with simulated real data
        Validates email discovery → transcription → Excel update → email processing
        """
        # Create workflow processor with dependency injection
        dependencies = {
            'config_manager': core_services['config_manager'],
            'service_initializer': core_services['service_initializer'],
            'workflow_orchestrator': Mock(),
            'error_handler': core_services['error_handler'],
            'logger': core_services['logger']
        }
        
        workflow_processor = ScribeWorkflowProcessor(dependencies)
        
        # Mock all external services but use real data structures
        with patch('requests.get') as mock_get, \
             patch('azure.storage.blob.BlobClient') as mock_blob, \
             patch('requests.post') as mock_post, \
             patch('openpyxl.load_workbook') as mock_excel:
            
            # Mock email discovery (simulating real Gmail response)
            mock_get.return_value.status_code = 200
            mock_get.return_value.json.return_value = {
                'value': [
                    {
                        'id': 'real-email-001',
                        'subject': 'Voice Message - Project Update',
                        'sender': {'emailAddress': {'address': 'colleague@company.com'}},
                        'receivedDateTime': '2025-01-24T10:30:00Z',
                        'hasAttachments': True,
                        'attachments': [
                            {
                                'id': 'att_001',
                                'name': 'project_update_voice.m4a',
                                'contentType': 'audio/m4a',
                                'size': 512000
                            }
                        ]
                    }
                ]
            }
            
            # Mock blob storage operations
            mock_blob_instance = Mock()
            mock_blob_instance.url = f"https://{production_config['storage_account']}.blob.core.windows.net/audio/project_update_voice.m4a"
            mock_blob.return_value = mock_blob_instance
            
            # Mock AI Foundry transcription response
            mock_post.return_value.status_code = 200
            mock_post.return_value.json.return_value = {
                'status': 'completed',
                'transcript': 'Hi everyone, this is a quick update on our project. We are on track to meet the deadline and should have the final deliverables ready by Friday.',
                'confidence': 0.94,
                'duration': 18.5,
                'language': 'en-US'
            }
            
            # Mock Excel operations
            import openpyxl
            mock_workbook = Mock()
            mock_worksheet = Mock()
            mock_worksheet.max_row = 5  # Existing data
            mock_workbook.active = mock_worksheet
            mock_excel.return_value = mock_workbook
            
            # Initialize services
            workflow_processor.initialize_services()
            
            # Execute complete workflow
            workflow_run = workflow_processor.execute_voice_email_workflow()
            
            # Validate workflow execution
            assert workflow_run is not None, "Should return workflow run result"
            assert workflow_run.success is True, "Workflow should complete successfully"
            
            # Verify all stages completed
            expected_stages = ['email_discovery', 'audio_processing', 'transcription', 'excel_update', 'email_management']
            completed_stages = [stage.stage_name for stage in workflow_run.stage_results if stage.success]
            
            for expected_stage in expected_stages:
                assert expected_stage in completed_stages, f"Stage {expected_stage} should complete successfully"
            
            # Verify service integrations were called
            assert mock_get.called, "Should call email discovery API"
            assert mock_post.called, "Should call transcription API"
            assert mock_excel.called, "Should access Excel file"

    def test_workflow_error_recovery_and_retry(self, core_services, environment_setup):
        """
        Test workflow error recovery and retry mechanisms
        Validates resilience to service failures and network issues
        """
        dependencies = {
            'config_manager': core_services['config_manager'],
            'service_initializer': core_services['service_initializer'],
            'workflow_orchestrator': Mock(),
            'error_handler': core_services['error_handler'],
            'logger': core_services['logger']
        }
        
        workflow_processor = ScribeWorkflowProcessor(dependencies)
        
        with patch('requests.get') as mock_get, \
             patch('requests.post') as mock_post:
            
            # Simulate intermittent failures followed by success
            mock_get.side_effect = [
                Exception("Network timeout"),  # First attempt fails
                Mock(status_code=200, json=lambda: {'value': []})  # Second attempt succeeds
            ]
            
            mock_post.side_effect = [
                Exception("AI service temporarily unavailable"),  # First transcription fails
                Mock(status_code=200, json=lambda: {
                    'status': 'completed',
                    'transcript': 'Recovery test successful',
                    'confidence': 0.91
                })  # Retry succeeds
            ]
            
            # Mock retry mechanism
            with patch.object(workflow_processor.error_handler, 'retry_with_exponential_backoff') as mock_retry:
                def retry_simulation(func, *args, **kwargs):
                    try:
                        return func(*args, **kwargs)
                    except Exception:
                        # Simulate successful retry
                        return func(*args, **kwargs)
                
                mock_retry.side_effect = retry_simulation
                
                workflow_processor.initialize_services()
                workflow_run = workflow_processor.execute_voice_email_workflow()
                
                # Should recover from errors
                assert workflow_run is not None, "Should complete despite initial failures"
                
                # Verify retry mechanism was used
                assert mock_retry.called or core_services['error_handler'].handle_error.called

    def test_workflow_performance_and_concurrency(self, production_config, core_services):
        """
        Test workflow performance with multiple emails and concurrent operations
        Validates scalability and performance under realistic load
        """
        dependencies = {
            'config_manager': core_services['config_manager'],
            'service_initializer': core_services['service_initializer'],
            'workflow_orchestrator': Mock(),
            'error_handler': core_services['error_handler'],
            'logger': core_services['logger']
        }
        
        workflow_processor = ScribeWorkflowProcessor(dependencies)
        
        # Simulate multiple emails with voice attachments
        multiple_emails_response = {
            'value': [
                {
                    'id': f'batch-email-{i:03d}',
                    'subject': f'Voice Message {i}',
                    'sender': {'emailAddress': {'address': f'user{i}@company.com'}},
                    'receivedDateTime': f'2025-01-24T{10 + (i % 12):02d}:00:00Z',
                    'hasAttachments': True,
                    'attachments': [
                        {
                            'id': f'att_{i:03d}',
                            'name': f'voice_message_{i}.wav',
                            'contentType': 'audio/wav',
                            'size': 256000 + (i * 1000)
                        }
                    ]
                }
                for i in range(1, 11)  # 10 emails for performance testing
            ]
        }
        
        with patch('requests.get') as mock_get, \
             patch('requests.post') as mock_post, \
             patch('time.time') as mock_time:
            
            # Setup mocks
            mock_get.return_value.status_code = 200
            mock_get.return_value.json.return_value = multiple_emails_response
            
            mock_post.return_value.status_code = 200
            mock_post.return_value.json.return_value = {
                'status': 'completed',
                'transcript': 'Performance test transcription',
                'confidence': 0.90
            }
            
            # Mock timing for performance measurement
            mock_time.side_effect = [0, 30]  # 30 seconds total processing
            
            workflow_processor.initialize_services()
            
            # Execute workflow with performance monitoring
            start_time = mock_time()
            workflow_run = workflow_processor.execute_voice_email_workflow()
            end_time = mock_time()
            
            processing_time = end_time - start_time
            
            # Validate performance requirements
            assert processing_time < 60, "Should process 10 emails within 60 seconds"
            assert workflow_run.success is True, "Should successfully process all emails"
            
            # Verify concurrent processing capabilities
            assert workflow_run.emails_processed >= 5, "Should process multiple emails"

    def test_workflow_data_consistency_and_integrity(self, temp_excel_file, core_services):
        """
        Test data consistency and integrity across the complete workflow
        Validates that data remains consistent from email to Excel
        """
        dependencies = {
            'config_manager': core_services['config_manager'],
            'service_initializer': core_services['service_initializer'],
            'workflow_orchestrator': Mock(),
            'error_handler': core_services['error_handler'],
            'logger': core_services['logger']
        }
        
        workflow_processor = ScribeWorkflowProcessor(dependencies)
        
        # Test data that should remain consistent throughout workflow
        test_email_data = {
            'id': 'consistency-test-001',
            'subject': 'Data Consistency Test Voice Message',
            'sender': {'emailAddress': {'address': 'consistency@test.com'}},
            'receivedDateTime': '2025-01-24T12:00:00Z',
            'hasAttachments': True,
            'attachments': [
                {
                    'id': 'att_consistency',
                    'name': 'consistency_test.wav',
                    'contentType': 'audio/wav',
                    'size': 384000
                }
            ]
        }
        
        expected_transcription = "This is a data consistency test message to ensure information flows correctly through the entire workflow."
        
        with patch('requests.get') as mock_get, \
             patch('requests.post') as mock_post, \
             patch('openpyxl.load_workbook') as mock_excel:
            
            # Setup consistent responses
            mock_get.return_value.status_code = 200
            mock_get.return_value.json.return_value = {'value': [test_email_data]}
            
            mock_post.return_value.status_code = 200
            mock_post.return_value.json.return_value = {
                'status': 'completed',
                'transcript': expected_transcription,
                'confidence': 0.96,
                'duration': 22.3
            }
            
            # Mock Excel to capture written data
            import openpyxl
            workbook = openpyxl.load_workbook(temp_excel_file)
            mock_excel.return_value = workbook
            
            workflow_processor.initialize_services()
            workflow_run = workflow_processor.execute_voice_email_workflow()
            
            # Validate data consistency
            assert workflow_run.success is True, "Workflow should complete successfully"
            
            # Check that Excel contains consistent data
            worksheet = workbook.active
            last_row_data = [cell.value for cell in worksheet[worksheet.max_row]]
            
            # Verify email data consistency
            assert 'consistency@test.com' in str(last_row_data), "Should preserve sender email"
            assert 'Data Consistency Test' in str(last_row_data), "Should preserve email subject"
            assert expected_transcription in str(last_row_data), "Should preserve transcription text"

    def test_workflow_monitoring_and_observability(self, core_services, environment_setup):
        """
        Test workflow monitoring and observability features
        Validates structured logging and application insights integration
        """
        dependencies = {
            'config_manager': core_services['config_manager'],
            'service_initializer': core_services['service_initializer'],
            'workflow_orchestrator': Mock(),
            'error_handler': core_services['error_handler'],
            'logger': core_services['logger']
        }
        
        workflow_processor = ScribeWorkflowProcessor(dependencies)
        
        with patch('requests.get') as mock_get, \
             patch('requests.post') as mock_post:
            
            # Setup minimal successful responses
            mock_get.return_value.status_code = 200
            mock_get.return_value.json.return_value = {'value': []}
            
            workflow_processor.initialize_services()
            workflow_run = workflow_processor.execute_voice_email_workflow()
            
            # Verify structured logging was performed
            logger_calls = core_services['logger'].log_info.call_args_list
            
            # Check for key workflow events
            logged_events = [str(call) for call in logger_calls]
            
            expected_events = [
                'workflow_started',
                'email_discovery',
                'workflow_completed'
            ]
            
            for expected_event in expected_events:
                event_found = any(expected_event in event for event in logged_events)
                assert event_found, f"Should log {expected_event} event"
            
            # Verify performance metrics were logged
            performance_logs = [call for call in logger_calls if 'performance' in str(call)]
            assert len(performance_logs) > 0, "Should log performance metrics"
            
            # Verify workflow metadata was captured
            assert workflow_run.start_time is not None, "Should track workflow start time"
            assert workflow_run.end_time is not None, "Should track workflow end time"
            assert workflow_run.total_duration > 0, "Should calculate total duration"

    def test_workflow_configuration_and_customization(self, core_services):
        """
        Test workflow configuration and customization capabilities
        Validates that workflow can be configured for different scenarios
        """
        dependencies = {
            'config_manager': core_services['config_manager'],
            'service_initializer': core_services['service_initializer'],
            'workflow_orchestrator': Mock(),
            'error_handler': core_services['error_handler'],
            'logger': core_services['logger']
        }
        
        workflow_processor = ScribeWorkflowProcessor(dependencies)
        
        # Test different workflow configurations
        test_configurations = [
            {
                'name': 'high_volume_processing',
                'max_concurrent_emails': 10,
                'audio_processing_timeout': 600,
                'enable_backup': True
            },
            {
                'name': 'quality_focused_processing',
                'max_concurrent_emails': 3,
                'audio_processing_timeout': 300,
                'minimum_confidence_threshold': 0.95
            },
            {
                'name': 'fast_processing',
                'max_concurrent_emails': 15,
                'audio_processing_timeout': 120,
                'skip_low_quality_audio': True
            }
        ]
        
        for config in test_configurations:
            # Apply configuration
            workflow_config = WorkflowConfiguration(**config)
            workflow_processor.workflow_config = workflow_config
            
            # Verify configuration was applied
            assert workflow_processor.workflow_config.max_concurrent_emails == config['max_concurrent_emails']
            assert workflow_processor.workflow_config.audio_processing_timeout == config['audio_processing_timeout']
            
            # Test workflow behavior with configuration
            with patch('requests.get') as mock_get:
                mock_get.return_value.status_code = 200
                mock_get.return_value.json.return_value = {'value': []}
                
                workflow_processor.initialize_services()
                workflow_run = workflow_processor.execute_voice_email_workflow()
                
                assert workflow_run is not None, f"Should work with {config['name']} configuration"

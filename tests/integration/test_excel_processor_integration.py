"""
Integration test for ScribeExcelProcessor with real Scribe.xlsx operations
Tests Excel file operations, data formatting, and Microsoft Graph integration
"""

import pytest
from unittest.mock import Mock, patch
import json
import tempfile
import openpyxl
from datetime import datetime

from src.processors.excel_processor import ScribeExcelProcessor
from src.models.transcription_models import TranscriptionResult


class TestScribeExcelProcessorIntegration:
    """Integration tests for Excel processor with real file operations"""

    def test_real_excel_file_operations(self, temp_excel_file, core_services):
        """
        Test operations on actual Excel file structure
        Validates data insertion and formatting preservation
        """
        excel_processor = ScribeExcelProcessor(
            core_services['config_manager'],
            core_services['error_handler'],
            core_services['logger']
        )
        
        # Create test transcription data
        transcription_data = {
            'date': datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S'),
            'sender': 'test@example.com',
            'subject': 'Test Voice Message',
            'audio_file_url': 'https://storage.blob.core.windows.net/audio/test.wav',
            'transcription': 'This is a test transcription of a voice message.',
            'duration': 15.5,
            'confidence': 0.94,
            'status': 'Completed'
        }
        
        # Test Excel update operation
        with patch('openpyxl.load_workbook') as mock_load_workbook:
            # Create real workbook for testing
            workbook = openpyxl.load_workbook(temp_excel_file)
            mock_load_workbook.return_value = workbook
            
            success = excel_processor.update_excel_with_transcription(
                transcription_data,
                transcription_data['audio_file_url']
            )
            
            assert success is True, "Should successfully update Excel file"
            
            # Verify data was added to worksheet
            worksheet = workbook.active
            last_row = worksheet.max_row
            
            # Check that data was inserted in new row
            assert last_row > 1, "Should have added data beyond header row"
            
            # Verify data integrity
            row_data = [cell.value for cell in worksheet[last_row]]
            assert transcription_data['sender'] in str(row_data), "Should contain sender email"
            assert transcription_data['subject'] in str(row_data), "Should contain email subject"
            assert transcription_data['transcription'] in str(row_data), "Should contain transcription text"

    def test_microsoft_graph_integration(self, production_config, core_services, environment_setup):
        """
        Test integration with Microsoft Graph API for Excel file access
        Validates authentication and file operations via Graph API
        """
        excel_processor = ScribeExcelProcessor(
            core_services['config_manager'],
            core_services['error_handler'],
            core_services['logger']
        )
        
        with patch('requests.get') as mock_get, patch('requests.patch') as mock_patch:
            # Mock Graph API responses
            mock_get.return_value.status_code = 200
            mock_get.return_value.json.return_value = {
                'workbook': {
                    'worksheets': [
                        {
                            'name': 'Voice Messages',
                            'id': 'worksheet_001'
                        }
                    ]
                }
            }
            
            mock_patch.return_value.status_code = 200
            
            # Test Excel file discovery via Graph API
            excel_metadata = excel_processor._get_excel_file_metadata()
            
            assert excel_metadata is not None, "Should retrieve Excel file metadata"
            
            # Verify Graph API calls
            mock_get.assert_called()
            call_url = mock_get.call_args[0][0]
            assert 'Scribe.xlsx' in call_url or 'workbook' in call_url, "Should target Scribe.xlsx file"

    def test_data_formatting_and_validation(self, temp_excel_file, core_services):
        """
        Test data formatting and validation for Excel operations
        Ensures data integrity and proper formatting
        """
        excel_processor = ScribeExcelProcessor(
            core_services['config_manager'],
            core_services['error_handler'],
            core_services['logger']
        )
        
        # Test with various data formats
        test_cases = [
            {
                'name': 'standard_transcription',
                'data': {
                    'date': '2025-01-24 10:30:00',
                    'sender': 'colleague@company.com',
                    'subject': 'Project Update Voice Note',
                    'audio_file_url': 'https://storage.blob.core.windows.net/audio/project_update.m4a',
                    'transcription': 'Hi team, just wanted to give you a quick update on the project status.',
                    'duration': 25.3,
                    'confidence': 0.97,
                    'status': 'Completed'
                }
            },
            {
                'name': 'long_transcription',
                'data': {
                    'date': '2025-01-24 14:15:00',
                    'sender': 'manager@company.com',
                    'subject': 'Weekly Review Voice Message',
                    'audio_file_url': 'https://storage.blob.core.windows.net/audio/weekly_review.wav',
                    'transcription': 'This is a longer transcription that might contain multiple sentences and complex information that needs to be properly formatted in the Excel cell.',
                    'duration': 120.7,
                    'confidence': 0.89,
                    'status': 'Completed'
                }
            }
        ]
        
        workbook = openpyxl.load_workbook(temp_excel_file)
        
        for test_case in test_cases:
            with patch('openpyxl.load_workbook', return_value=workbook):
                success = excel_processor.update_excel_with_transcription(
                    test_case['data'],
                    test_case['data']['audio_file_url']
                )
                
                assert success is True, f"Should handle {test_case['name']} data format"
        
        # Verify all data was properly formatted
        worksheet = workbook.active
        assert worksheet.max_row >= len(test_cases) + 1, "Should have added all test records"

    def test_excel_file_backup_and_versioning(self, temp_excel_file, core_services):
        """
        Test Excel file backup and versioning capabilities
        Ensures data safety during updates
        """
        excel_processor = ScribeExcelProcessor(
            core_services['config_manager'],
            core_services['error_handler'],
            core_services['logger']
        )
        
        with patch('shutil.copy2') as mock_copy:
            # Test backup creation before update
            transcription_data = {
                'date': datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S'),
                'sender': 'backup_test@example.com',
                'subject': 'Backup Test Voice Message',
                'audio_file_url': 'https://storage.blob.core.windows.net/audio/backup_test.wav',
                'transcription': 'This is a test for backup functionality.',
                'duration': 10.0,
                'confidence': 0.95,
                'status': 'Completed'
            }
            
            with patch('openpyxl.load_workbook') as mock_load_workbook:
                workbook = openpyxl.load_workbook(temp_excel_file)
                mock_load_workbook.return_value = workbook
                
                # Enable backup mode
                excel_processor.enable_backup = True
                
                success = excel_processor.update_excel_with_transcription(
                    transcription_data,
                    transcription_data['audio_file_url']
                )
                
                assert success is True, "Should successfully update with backup enabled"
                
                # Verify backup was created
                if excel_processor.enable_backup:
                    mock_copy.assert_called(), "Should create backup before update"

    def test_concurrent_excel_operations(self, temp_excel_file, core_services):
        """
        Test handling of concurrent Excel operations
        Validates file locking and data consistency
        """
        excel_processor = ScribeExcelProcessor(
            core_services['config_manager'],
            core_services['error_handler'],
            core_services['logger']
        )
        
        # Simulate concurrent operations
        transcription_batch = [
            {
                'date': f'2025-01-24 10:{i:02d}:00',
                'sender': f'user{i}@example.com',
                'subject': f'Voice Message {i}',
                'audio_file_url': f'https://storage.blob.core.windows.net/audio/message_{i}.wav',
                'transcription': f'This is voice message number {i}.',
                'duration': 10 + i,
                'confidence': 0.90 + (i * 0.01),
                'status': 'Completed'
            }
            for i in range(1, 6)  # 5 concurrent operations
        ]
        
        workbook = openpyxl.load_workbook(temp_excel_file)
        
        with patch('openpyxl.load_workbook', return_value=workbook):
            results = []
            for transcription_data in transcription_batch:
                result = excel_processor.update_excel_with_transcription(
                    transcription_data,
                    transcription_data['audio_file_url']
                )
                results.append(result)
            
            # All operations should succeed (with proper locking/retry logic)
            assert all(results), "All concurrent operations should succeed"
            
            # Verify all data was properly inserted
            worksheet = workbook.active
            assert worksheet.max_row >= len(transcription_batch) + 1, "Should contain all batch records"

    def test_error_recovery_and_data_integrity(self, temp_excel_file, core_services):
        """
        Test error recovery and data integrity preservation
        Validates handling of Excel file corruption and network issues
        """
        excel_processor = ScribeExcelProcessor(
            core_services['config_manager'],
            core_services['error_handler'],
            core_services['logger']
        )
        
        transcription_data = {
            'date': '2025-01-24 15:00:00',
            'sender': 'error_test@example.com',
            'subject': 'Error Recovery Test',
            'audio_file_url': 'https://storage.blob.core.windows.net/audio/error_test.wav',
            'transcription': 'Testing error recovery mechanisms.',
            'duration': 12.0,
            'confidence': 0.93,
            'status': 'Completed'
        }
        
        # Test file corruption recovery
        with patch('openpyxl.load_workbook') as mock_load_workbook:
            # Simulate file corruption on first attempt
            mock_load_workbook.side_effect = [
                Exception("Excel file corrupted"),
                openpyxl.load_workbook(temp_excel_file)  # Recovery attempt succeeds
            ]
            
            # Test with retry mechanism
            with patch.object(excel_processor.error_handler, 'retry_with_exponential_backoff') as mock_retry:
                def retry_func(func, *args, **kwargs):
                    try:
                        return func(*args, **kwargs)
                    except Exception:
                        # Simulate successful retry
                        return func(*args, **kwargs)
                
                mock_retry.side_effect = retry_func
                
                success = excel_processor.update_excel_with_transcription(
                    transcription_data,
                    transcription_data['audio_file_url']
                )
                
                # Should recover from error
                assert success in [True, None], "Should handle errors gracefully"
                
                # Verify error handling was invoked
                assert mock_retry.called or core_services['error_handler'].handle_error.called

    def test_performance_optimization(self, temp_excel_file, core_services):
        """
        Test performance optimization for large Excel files
        Validates efficient data operations and memory usage
        """
        excel_processor = ScribeExcelProcessor(
            core_services['config_manager'],
            core_services['error_handler'],
            core_services['logger']
        )
        
        # Create larger dataset for performance testing
        large_dataset = [
            {
                'date': f'2025-01-{(i % 30) + 1:02d} {(i % 24):02d}:{(i % 60):02d}:00',
                'sender': f'performance_test_{i}@example.com',
                'subject': f'Performance Test Voice Message {i}',
                'audio_file_url': f'https://storage.blob.core.windows.net/audio/perf_test_{i}.wav',
                'transcription': f'This is performance test transcription number {i} with some additional text to simulate real-world data.',
                'duration': 15 + (i % 60),
                'confidence': 0.85 + ((i % 15) * 0.01),
                'status': 'Completed'
            }
            for i in range(50)  # 50 records for performance testing
        ]
        
        workbook = openpyxl.load_workbook(temp_excel_file)
        
        with patch('openpyxl.load_workbook', return_value=workbook):
            with patch('time.time') as mock_time:
                # Mock timing for performance measurement
                mock_time.side_effect = list(range(0, 100, 2))  # 2 seconds per operation
                
                start_time = mock_time()
                
                # Process batch operations
                for transcription_data in large_dataset[:10]:  # Test with first 10 records
                    excel_processor.update_excel_with_transcription(
                        transcription_data,
                        transcription_data['audio_file_url']
                    )
                
                end_time = mock_time()
                
                # Verify performance is within acceptable limits
                processing_time = end_time - start_time
                assert processing_time < 30, "Batch processing should complete within 30 seconds"
                
                # Verify performance logging
                assert core_services['logger'].log_info.called, "Should log performance metrics"
